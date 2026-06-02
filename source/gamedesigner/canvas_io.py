from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path

from .data_canvas import layout_canvas_nodes
from .models import BlueprintGroup, CanvasData, FIELD_TYPES, NodeField, NodeTemplate, ProjectData, new_id


SHEET_TYPE_VALUES = set(FIELD_TYPES) | {"数字"}
NODE_LAYOUT_IMPORT_HEADERS = {
    "x": {"节点X", "node_x"},
    "y": {"节点Y", "node_y"},
}
GROUP_IMPORT_HEADERS = {
    "id": {"蓝图组ID", "group_id", "blueprint_group_id"},
    "title": {"蓝图组名称", "蓝图组", "group_title", "blueprint_group"},
    "x": {"蓝图组X", "group_x"},
    "y": {"蓝图组Y", "group_y"},
    "width": {"蓝图组宽", "group_width"},
    "height": {"蓝图组高", "group_height"},
    "color": {"蓝图组颜色", "group_color"},
}
GROUP_PADDING_X = 40.0
GROUP_PADDING_TOP = 56.0
GROUP_PADDING_BOTTOM = 40.0
DEFAULT_IMPORTED_NODE_WIDTH = 360.0
DEFAULT_IMPORTED_NODE_HEIGHT = 220.0


@dataclass(frozen=True)
class ImportedSheet:
    headers: list[str]
    rows: list[list[str]]


def import_canvas_sheet(
    project: ProjectData,
    canvas: CanvasData,
    source: str | Path,
) -> NodeTemplate:
    sheet = read_sheet(source)
    group_columns = _group_metadata_columns(sheet.headers)
    node_layout_columns = _node_layout_columns(sheet.headers)
    metadata_indexes = set(group_columns.values()) | set(node_layout_columns.values())
    field_indexes = [index for index in range(len(sheet.headers)) if index not in metadata_indexes]
    template_headers = [sheet.headers[index] for index in field_indexes]
    template = _build_template_for_canvas(canvas, template_headers)
    _replace_or_add_template(project, template)
    canvas.template_id = template.id
    canvas.nodes.clear()
    canvas.edges.clear()
    canvas.groups.clear()
    imported_groups: dict[str, BlueprintGroup] = {}
    for index, row in enumerate(sheet.rows, start=1):
        node = template.create_node(0.0, 0.0)
        node.template_locked = canvas.is_data_canvas()
        for field_index, field in enumerate(node.fields):
            source_index = field_indexes[field_index]
            field.value = _cell(row, source_index)
        title_field = next((field for field in node.fields if field.id == node.title_field_id), None)
        if title_field is not None:
            title = title_field.value.strip() or title_field.name.strip() or template.name
            node.title = title
        _apply_node_layout(row, node_layout_columns, node)
        group = _group_for_import_row(row, group_columns, imported_groups)
        if group is not None:
            node.group_id = group.id
        node.order = index
        canvas.nodes.append(node)
    canvas.groups.extend(imported_groups.values())
    if not node_layout_columns:
        layout_canvas_nodes(canvas)
    if not _has_group_layout_columns(group_columns):
        _layout_imported_groups(canvas)
    return template


def read_sheet(source: str | Path) -> ImportedSheet:
    path = Path(source)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_excel(path)
    raise ValueError("仅支持导入 CSV 或 Excel（.xlsx/.xlsm）文件。")


def _read_csv(path: Path) -> ImportedSheet:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    return _normalize_rows(rows)


def _read_excel(path: Path) -> ImportedSheet:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("缺少 Excel 导入依赖 openpyxl，请重新安装依赖或重新打包 exe。") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for values in sheet.iter_rows(values_only=True):
        rows.append(["" if value is None else str(value) for value in values])
    workbook.close()
    return _normalize_rows(rows)


def _normalize_rows(raw_rows: list[list[str]]) -> ImportedSheet:
    trimmed = [_trim_row(list(row)) for row in raw_rows]
    meaningful = [row for row in trimmed if any(cell.strip() for cell in row)]
    if not meaningful:
        raise ValueError("表格内容为空。")
    headers = meaningful[0]
    if not any(header.strip() for header in headers):
        raise ValueError("表头不能为空。")
    normalized_headers = [_normalize_header(header, index) for index, header in enumerate(headers, start=1)]
    width = len(normalized_headers)
    data_rows = meaningful[1:]
    if data_rows and _looks_like_type_row(_pad_row(data_rows[0], width)):
        data_rows = data_rows[1:]
    rows = [_pad_row(row, width) for row in data_rows]
    return ImportedSheet(headers=normalized_headers, rows=rows)


def _build_template_for_canvas(canvas: CanvasData, headers: list[str]) -> NodeTemplate:
    fields = [NodeField(name=header, data_type="文本", value="") for header in headers]
    title_field_id = fields[0].id if fields else ""
    return NodeTemplate(
        id=canvas.template_id or new_id("template"),
        name=f"{canvas.name} 模板",
        icon="数" if canvas.is_data_canvas() else "N",
        icon_from_title=not canvas.is_data_canvas(),
        title_field_id=title_field_id,
        fields=fields,
    )


def _replace_or_add_template(project: ProjectData, template: NodeTemplate) -> None:
    for index, existing in enumerate(project.templates):
        if existing.id == template.id:
            project.templates[index] = template
            return
    project.templates.append(template)


def _group_metadata_columns(headers: list[str]) -> dict[str, int]:
    has_group_id = any(header.strip() in GROUP_IMPORT_HEADERS["id"] for header in headers)
    if not has_group_id:
        return {}
    columns: dict[str, int] = {}
    for index, header in enumerate(headers):
        name = header.strip()
        for prop, names in GROUP_IMPORT_HEADERS.items():
            if prop not in columns and name in names:
                columns[prop] = index
                break
    return columns


def _node_layout_columns(headers: list[str]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for index, header in enumerate(headers):
        name = header.strip()
        for prop, names in NODE_LAYOUT_IMPORT_HEADERS.items():
            if prop not in columns and name in names:
                columns[prop] = index
                break
    return columns


def _group_for_import_row(
    row: list[str],
    group_columns: dict[str, int],
    imported_groups: dict[str, BlueprintGroup],
) -> BlueprintGroup | None:
    if not group_columns:
        return None
    group_id = _cell(row, group_columns.get("id")).strip()
    title = _cell(row, group_columns.get("title")).strip()
    if not group_id and not title:
        return None
    key = group_id or f"title:{title.casefold()}"
    group = imported_groups.get(key)
    if group is None:
        group = BlueprintGroup(
            id=group_id or new_id("group"),
            title=title or "蓝图组",
        )
        imported_groups[key] = group
    elif title and group.title == "蓝图组":
        group.title = title
    _apply_group_layout(row, group_columns, group)
    return group


def _apply_node_layout(row: list[str], node_layout_columns: dict[str, int], node: object) -> None:
    x = _float_or_none(_cell(row, node_layout_columns.get("x")))
    y = _float_or_none(_cell(row, node_layout_columns.get("y")))
    if x is not None:
        setattr(node, "x", x)
    if y is not None:
        setattr(node, "y", y)


def _apply_group_layout(row: list[str], group_columns: dict[str, int], group: BlueprintGroup) -> None:
    x = _float_or_none(_cell(row, group_columns.get("x")))
    y = _float_or_none(_cell(row, group_columns.get("y")))
    width = _float_or_none(_cell(row, group_columns.get("width")))
    height = _float_or_none(_cell(row, group_columns.get("height")))
    color = _cell(row, group_columns.get("color")).strip()
    if x is not None:
        group.x = x
    if y is not None:
        group.y = y
    if width is not None:
        group.width = max(180.0, width)
    if height is not None:
        group.height = max(120.0, height)
    if color:
        group.color = color


def _has_group_layout_columns(group_columns: dict[str, int]) -> bool:
    return any(prop in group_columns for prop in ("x", "y", "width", "height"))


def _layout_imported_groups(canvas: CanvasData) -> None:
    if not canvas.groups:
        return
    for group in canvas.groups:
        members = [node for node in canvas.nodes if node.group_id == group.id]
        if not members:
            continue
        min_x = min(node.x for node in members)
        min_y = min(node.y for node in members)
        max_x = max(node.x + _node_width(node) for node in members)
        max_y = max(node.y + _node_height(node) for node in members)
        group.x = min_x - GROUP_PADDING_X
        group.y = min_y - GROUP_PADDING_TOP
        group.width = max(180.0, max_x - min_x + GROUP_PADDING_X * 2.0)
        group.height = max(120.0, max_y - min_y + GROUP_PADDING_TOP + GROUP_PADDING_BOTTOM)


def _cell(row: list[str], index: int | None) -> str:
    if index is None or index < 0 or index >= len(row):
        return ""
    return row[index]


def _float_or_none(value: str) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _node_width(node: object) -> float:
    width = float(getattr(node, "width", 0.0) or 0.0)
    return width if width > 0 else DEFAULT_IMPORTED_NODE_WIDTH


def _node_height(node: object) -> float:
    height = float(getattr(node, "height", 0.0) or 0.0)
    return height if height > 0 else DEFAULT_IMPORTED_NODE_HEIGHT


def _trim_row(row: list[str]) -> list[str]:
    end = len(row)
    while end > 0 and not str(row[end - 1]).strip():
        end -= 1
    return [str(cell) for cell in row[:end]]


def _pad_row(row: list[str], width: int) -> list[str]:
    if len(row) < width:
        return [*row, *([""] * (width - len(row)))]
    return row[:width]


def _looks_like_type_row(row: list[str]) -> bool:
    nonempty = [cell.strip() for cell in row if cell.strip()]
    return bool(nonempty) and all(cell in SHEET_TYPE_VALUES for cell in nonempty)


def _normalize_header(value: str, index: int) -> str:
    text = value.strip()
    return text or f"字段{index}"
